"""Parse Hostname/IP inventory files (.xlsx / .csv) for bulk server import."""

from __future__ import annotations

import csv
import io
import re
from typing import Any

HOSTNAME_ALIASES = {"hostname", "host", "fqdn", "server", "sunucu", "name"}
IP_ALIASES = {"ip", "ip_address", "ipaddress", "address", "adres"}


def _norm_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("\ufeff", "")
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")


def _pick_columns(headers: list[str]) -> tuple[int, int]:
    host_idx = -1
    ip_idx = -1
    for i, h in enumerate(headers):
        key = _norm_header(h)
        if host_idx < 0 and key in HOSTNAME_ALIASES:
            host_idx = i
        if ip_idx < 0 and key in IP_ALIASES:
            ip_idx = i
    if host_idx < 0 or ip_idx < 0:
        raise ValueError(
            "Dosyada Hostname ve IP sütunları gerekli "
            "(ör. Hostname, IP / host, ip_address)"
        )
    return host_idx, ip_idx


def _rows_from_matrix(matrix: list[list[Any]]) -> list[dict[str, str]]:
    if not matrix:
        raise ValueError("Dosya boş")
    # Find first non-empty row as header
    header_row_idx = 0
    while header_row_idx < len(matrix) and not any(
        str(c or "").strip() for c in matrix[header_row_idx]
    ):
        header_row_idx += 1
    if header_row_idx >= len(matrix):
        raise ValueError("Dosya boş")
    headers = [str(c or "") for c in matrix[header_row_idx]]
    host_idx, ip_idx = _pick_columns(headers)
    out: list[dict[str, str]] = []
    for row in matrix[header_row_idx + 1 :]:
        if not row:
            continue
        hostname = str(row[host_idx] if host_idx < len(row) else "").strip()
        ip = str(row[ip_idx] if ip_idx < len(row) else "").strip()
        if not hostname and not ip:
            continue
        if not hostname or not ip:
            raise ValueError(f"Eksik Hostname/IP satırı: hostname={hostname!r} ip={ip!r}")
        out.append({"hostname": hostname, "ip": ip})
    if not out:
        raise ValueError("İçe aktarılacak satır bulunamadı")
    return out


def parse_csv_bytes(data: bytes) -> list[dict[str, str]]:
    text = data.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    matrix = [list(row) for row in reader]
    return _rows_from_matrix(matrix)


def parse_xlsx_bytes(data: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ValueError("Excel desteği için openpyxl gerekli") from exc
    wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb.active
        matrix: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            matrix.append(list(row))
    finally:
        wb.close()
    return _rows_from_matrix(matrix)


def parse_inventory_file(filename: str, data: bytes) -> list[dict[str, str]]:
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return parse_xlsx_bytes(data)
    if name.endswith(".csv") or name.endswith(".txt"):
        return parse_csv_bytes(data)
    # sniff: zip/xlsx starts with PK
    if data[:2] == b"PK":
        return parse_xlsx_bytes(data)
    return parse_csv_bytes(data)
